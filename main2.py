from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook
import tkinter.messagebox as messagebox

root = Tk()
root.geometry('950x552')
root.title("Aroma Elegance")

now = datetime.datetime.now()
date = now.strftime("%Y-%m-%d")

# =========Data Exel===========#
wb = Workbook()
ws = wb.active

ws.title = 'customer'
ws["A1"] = 'Full Name'
ws["B1"] = 'Number Phone'
ws["C1"] = 'Address'
ws["D1"] = 'Total'
ws["E1"] = 'Date Bye'
wb.save('aroma.xlsx')


def save():
    name = En_name.get()
    phone = En_phone.get()
    address = En_address.get()
    total = En_total.get()
    datebay = En_date.get()

    excel = openpyxl.load_workbook('aroma.xlsx')
    file = excel.active
    file.cell(column=1, row=file.max_row + 1, value=name)
    file.cell(column=2, row=file.max_row, value=phone)
    file.cell(column=3, row=file.max_row, value=address)
    file.cell(column=4, row=file.max_row, value=total)
    file.cell(column=5, row=file.max_row, value=datebay)
    excel.save('aroma.xlsx')


def search():
    name = En_name.get()
    excel = openpyxl.load_workbook('aroma.xlsx')
    file = excel.active

    for row in file.iter_rows(min_row=2, values_only=True):
        if row[0] == name:
            En_phone.delete(0, END)
            En_phone.insert(0, row[1])
            En_address.delete(0, END)
            En_address.insert(0, row[2])
            En_total.delete(0, END)
            En_total.insert(0, row[3])
            En_date.delete(0, END)
            En_date.insert(0, row[4])
            messagebox.showinfo("Search Result", "Record found.")
            return

    messagebox.showerror("Search Result", "Record not found.")


def delete():
    name = En_name.get()
    excel = openpyxl.load_workbook('aroma.xlsx')
    file = excel.active

    for row in file.iter_rows(min_row=2):
        if row[0].value == name:
            file.delete_rows(row[0].row, 1)
            excel.save('aroma.xlsx')
            clear()
            messagebox.showinfo("Delete Result", "Record deleted.")
            return

    messagebox.showerror("Delete Result", "Record not found.")


def clear():
    En_name.delete(0, END)
    En_phone.delete(0, END)
    En_address.delete(0, END)
    En_total.delete(0, END)
    En_date.delete(0, END)
    
    for item in trv.get_children():
        trv.delete(item)
    
    for i in range(12):
        sv[i].set(0)


def close_program():
    root.quit()


# UI elements and placement
Label(root, text='Full Name').pack()
En_name = Entry(root)
En_name.pack()

Label(root, text='Number Phone').pack()
En_phone = Entry(root)
En_phone.pack()

Label(root, text='Address').pack()
En_address = Entry(root)
En_address.pack()

Label(root, text='Total').pack()
En_total = Entry(root)
En_total.pack()

Label(root, text='Date Bye').pack()
En_date = Entry(root)
En_date.pack()

# Buttons
b1 = Button(root, text='Save', command=save)
b1.pack()

b2 = Button(root, text='Search', command=search)
b2.pack()

b3 = Button(root, text='Delete', command=delete)
b3.pack()

b4 = Button(root, text='New Invoice', command=clear)
b4.pack()

b5 = Button(root, text='Close', command=close_program)
b5.pack()

# إضافات جديدة بناءً على الكود السابق الذي قدمته
#======price======#
menu1 = {
    0: ['Our Moment', 50],
    1: ['Shalimar', 150],
    2: ['Si', 130],
    3: ['My Way', 250],
    4: ['Sauvage', 350],
    5: ['Valentino Uomo', 450],
    6: ['212 VIP', 550],
    7: ['Hugo Boss', 500],
    8: ['Boss', 100],
    9: ['Jadore', 200],
    10: ['Coco Chanel', 300],
    11: ['Prada', 400],
}

def billd():
    global En_name 
    global En_phone
    global En_address
    global En_total
    global En_date
    
    # root.geometry('1205x552')  # Remove this line to keep the window size fixed
    F4 = Frame(root, bg='#E5D4FF', width=250, height=544, bd=2, relief=GROOVE)
    F4.place(x=950, y=1)

    L_name = Label(F4, text='Buyer Name', bg='#E5D4FF', fg='black')
    L_name.place(x=168, y=10)
    En_name = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
    En_name.place(x=15, y=40)

    L_phone = Label(F4, text='Buyer Number', bg='#E5D4FF', fg='black')
    L_phone.place(x=170, y=70)
    En_phone = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
    En_phone.place(x=15, y=100)

    L_address = Label(F4, text='Buyer Address', bg='#E5D4FF', fg='black')
    L_address.place(x=160, y=130)
    En_address = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
    En_address.place(x=15, y=160)

    L_total = Label(F4, text='Total', bg='#E5D4FF', fg='black')
    L_total.place(x=165, y=190)
    En_total = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
    En_total.place(x=15, y=210)

    L_date = Label(F4, text='Date', bg='#E5D4FF', fg='black')
    L_date.place(x=175, y=240)
    En_date = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
    En_date.place(x=15, y=270)

    save_button = Button(F4, text='Save', width=31, cursor='hand2', bg='#FFCDEA', fg='#240750', height=2, command=save)
    save_button.place(x=12, y=310)

    empty_button = Button(F4, text='Empty', width=31, cursor='hand2', bg='#FFCDEA', fg='#240750', height=2, command=clear1)
    empty_button.place(x=12, y=370)

    search_button = Button(F4, text='Search', width=31, cursor='hand2', bg='#FFCDEA', fg='#240750', height=2, command=search)
    search_button.place(x=12, y=430)

    delete_button = Button(F4, text='Delete', width=31, cursor='hand2', bg='#FFCDEA', fg='#240750', height=2, command=delete)
    delete_button.place(x=12, y=490)

    total = 0
    # حذف جميع العناصر الحالية من Treeview
    for i in trv.get_children():
        trv.delete(i)
    for j in range(len(sb)):
        if int(sb[j].get()) > 0:
            price = int(sb[j].get()) * menu1[j][1]
            total += price
            m = (str(sb[j].get()), str(menu1[j][1]), str(price))
            trv.insert("", 'end', iid=j, text=menu1[j][0], values=m)
    finall = total
    En_total.insert('1', str(finall) + '$')
    En_date.insert('1', str(date))


def clear1():
    En_name.delete('0', END)
    En_phone.delete('0', END)
    En_address.delete('0', END)
    En_total.delete('0', END)
    En_date.delete('0', END)


#============= [Frame [1]] =======
F1 = Frame(root, bg='#E5D4FF', width=600, height=550)
F1.place(x=1, y=1)

#============= [Image] =======
img_menu1 = PhotoImage(file=r'img/1.png')
img_menu2 = PhotoImage(file=r'img/2.png')
img_menu3 = PhotoImage(file=r'img/3.png')
img_menu4 = PhotoImage(file=r'img/4.png')
img_menu5 = PhotoImage(file=r'img/5.png')
img_menu6 = PhotoImage(file=r'img/6.png')
img_menu7 = PhotoImage(file=r'img/7.png')
img_menu8 = PhotoImage(file=r'img/8.png')
img_menu9 = PhotoImage(file=r'img/9.png')
img_menu10 = PhotoImage(file=r'img/10.png')
img_menu11 = PhotoImage(file=r'img/11.png')
img_menu12 = PhotoImage(file=r'img/12.png')

# Variable to display the name of this frame
title = Label(F1, text="Welcome to the World of Elegance, Welcome to Aroma Elegance!", font=('Tajawal 13'), fg='white', bg='#BC7FCD', width=70)
title.place(x=0, y=0)

# Buttons for the first row
menue1 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu1, text='Our Moment', compound=TOP)
menue1.place(x=30, y=45)
menue2 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu2, text='Shalimar', compound=TOP)
menue2.place(x=170, y=45)
menue3 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu3, text='Si', compound=TOP)
menue3.place(x=310, y=45)
menue4 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu4, text='My Way', compound=TOP)
menue4.place(x=450, y=45)

# Buttons for the second row
menue5 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu5, text='Sauvage', compound=TOP)
menue5.place(x=30, y=180)
menue6 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu6, text='Valentino Uomo', compound=TOP)
menue6.place(x=170, y=180)
menue7 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu7, text='212 VIP', compound=TOP)
menue7.place(x=310, y=180)
menue8 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu8, text='Hugo Boss', compound=TOP)
menue8.place(x=450, y=180)

# Buttons for the third row
menue9 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu9, text='Boss', compound=TOP)
menue9.place(x=30, y=320)
menue10 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu10, text='Jadore', compound=TOP)
menue10.place(x=170, y=320)
menue11 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu11, text='Coco Chanel', compound=TOP)
menue11.place(x=310, y=320)
menue12 = Button(F1, width=88, bg='#C4DFDF', bd=1, relief=SOLID, cursor='hand2', height=85, image=img_menu12, text='Prada', compound=TOP)
menue12.place(x=450, y=320)

# Spinboxes


# Spinboxes
sb = []
font1 = ('Times', 12)
sv = [IntVar() for _ in range(12)]

# الصف الأول
for i in range(4):
    sb.append(Spinbox(F1, from_=0, to=5, font=font1, width=10, textvariable=sv[i]))
    sb[i].place(x=30 + i*140, y=140)

# الصف الثاني
for i in range(4, 8):
    sb.append(Spinbox(F1, from_=0, to=5, font=font1, width=10, textvariable=sv[i]))
    sb[i].place(x=30 + (i-4)*140, y=275)

# الصف الثالث
for i in range(8, 12):
    sb.append(Spinbox(F1, from_=0, to=5, font=font1, width=10, textvariable=sv[i]))
    sb[i].place(x=30 + (i-8)*140, y=415)

# Buttons
b1 = Button(F1, text='Buying', fg='#240750', font=('Tajawal 12'), width=15, bg='#FFCDEA', bd=1, relief=SOLID, cursor='hand2', height=2, command=billd)
b1.place(x=20, y=480)
b2 = Button(F1, text='New invoice', fg='#240750', font=('Tajawal 12'), width=15, bg='#FFCDEA', bd=1, relief=SOLID, cursor='hand2', height=2, command=clear)
b2.place(x=210, y=480)
b3 = Button(F1, text='Close the program', fg='#240750', font=('Tajawal 12'), width=15, bg='#FFCDEA', bd=1, relief=SOLID, cursor='hand2', height=2, command=close_program)
b3.place(x=400, y=480)

# Frame[2]
F2 = Frame(root, bg='gray', width=343, height=550)
F2.place(x=604, y=1)

trv = ttk.Treeview(F2, selectmode='browse')
trv.place(x=1, y=1, width=340, height=550)

trv["columns"] = ('1', '2', '3')
trv.column("#0", width=80, anchor='c')
trv.column("1", width=50, anchor='c')
trv.column("2", width=50, anchor='c')
trv.column("3", width=60, anchor='c')

trv.heading("#0", text="Product", anchor='c')
trv.heading("1", text="Number", anchor='c')
trv.heading("2", text="Price", anchor='c')
trv.heading("3", text="Total account", anchor='c')


root.mainloop()